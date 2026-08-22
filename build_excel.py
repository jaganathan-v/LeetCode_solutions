"""
build_excel.py
---------------
Scans every problem folder in this repo (as pushed here by LeetSync),
pulls the official topic tags straight from LeetCode's public API,
reads runtime/memory/date from the git commit history LeetSync already
created, and regenerates reports/LeetCode_Analytics.xlsx.

Runs automatically via .github/workflows/update-excel.yml on every push.
Can also be run locally: `python scripts/build_excel.py`
"""

import os
import re
import json
import subprocess
import time
from datetime import datetime
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(REPO_ROOT, "reports", "LeetCode_Analytics.xlsx")
CACHE_PATH = os.path.join(REPO_ROOT, "scripts", "tag_cache.json")
IGNORE_DIRS = {".git", ".github", "reports", "scripts", "node_modules"}

# Priority order used to pick a single "Topic" out of LeetCode's tag list.
# Whichever tag appears first in this list wins; everything else becomes Subtopic.
TOPIC_PRIORITY = [
    "Array", "String", "Tree", "Binary Tree", "Graph", "Stack", "Queue",
    "Linked List", "Dynamic Programming", "Math", "Hash Table",
    "Two Pointers", "Sliding Window", "Binary Search", "Greedy",
    "Backtracking", "Bit Manipulation", "Heap (Priority Queue)", "Trie",
    "Sorting", "Simulation", "Recursion", "Design", "Union Find",
]
TREE_ALIASES = {"Binary Tree", "Binary Search Tree", "Binary Tree Traversal"}


# ----------------------------------------------------------------------------
# 1. Fetch official topic tags from LeetCode's public GraphQL API
# ----------------------------------------------------------------------------
def load_cache():
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)


def fetch_tags(slug, cache):
    """Returns a list of tag names for a problem slug, e.g. ['Array','Hash Table']."""
    if slug in cache:
        return cache[slug]

    query = """
    query questionTags($titleSlug: String!) {
      question(titleSlug: $titleSlug) {
        topicTags { name }
      }
    }
    """
    try:
        resp = requests.post(
            "https://leetcode.com/graphql",
            json={"query": query, "variables": {"titleSlug": slug}},
            headers={
                "Content-Type": "application/json",
                "Referer": f"https://leetcode.com/problems/{slug}/",
                "User-Agent": "Mozilla/5.0 (analytics-bot)",
            },
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        tags = [t["name"] for t in data["data"]["question"]["topicTags"]]
        if not tags:
            tags = ["Uncategorized"]
    except Exception as e:
        print(f"  [warn] tag fetch failed for {slug}: {e}")
        tags = ["Uncategorized"]

    cache[slug] = tags
    save_cache(cache)
    time.sleep(0.4)  # be polite to LeetCode's endpoint
    return tags


def classify(tags):
    for pref in TOPIC_PRIORITY:
        if pref in tags:
            topic = "Tree" if pref in TREE_ALIASES else pref
            rest = [t for t in tags if t != pref]
            subtopic = ", ".join(rest) if rest else "General"
            return topic, subtopic
    topic = tags[0] if tags else "Other"
    subtopic = ", ".join(tags[1:]) if len(tags) > 1 else "General"
    return topic, subtopic


# ----------------------------------------------------------------------------
# 2. Parse each problem folder: README (title/difficulty/slug) + code file
# ----------------------------------------------------------------------------
def parse_readme(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        content = f.read()
    out = {}
    m = re.search(r'<h2><a href="([^"]+)">([^<]+)</a></h2>', content)
    if m:
        out["url"] = m.group(1)
        out["title"] = m.group(2)
        slug_m = re.search(r"/problems/([a-z0-9\-]+)/?", m.group(1))
        out["slug"] = slug_m.group(1) if slug_m else None
    d = re.search(r"Difficulty-(\w+)-", content)
    if d:
        out["difficulty"] = d.group(1)
    return out


def git_commit_stats(folder_name):
    """Pulls Time/Memory from LeetSync's own commit messages for this folder."""
    try:
        log = subprocess.run(
            ["git", "log", "--follow", "--date=iso-strict",
             "--format=%ad|%s", "--", folder_name],
            cwd=REPO_ROOT, capture_output=True, text=True, check=True,
        ).stdout.strip().splitlines()
    except Exception:
        return {}

    entries = []
    for line in log:
        if "|" not in line:
            continue
        date, msg = line.split("|", 1)
        m = re.match(
            r"Time: (\d+) ms \(([\d.]+)%\) \| Memory: ([\d.]+) MB \(([\d.]+)%\)", msg
        )
        if m:
            entries.append({
                "date": date,
                "runtime_ms": int(m.group(1)),
                "runtime_pct": float(m.group(2)),
                "memory_mb": float(m.group(3)),
                "memory_pct": float(m.group(4)),
            })
    if not entries:
        return {}
    entries.sort(key=lambda e: e["date"])  # oldest -> newest
    return {"first_date": entries[0]["date"], "latest": entries[-1], "attempts": len(entries)}


def collect_records():
    cache = load_cache()
    records = []
    for folder in sorted(os.listdir(REPO_ROOT)):
        fpath = os.path.join(REPO_ROOT, folder)
        if not os.path.isdir(fpath) or folder in IGNORE_DIRS or folder.startswith("."):
            continue
        files = os.listdir(fpath)
        readmes = [f for f in files if f.lower() == "readme.md"]
        code_files = [f for f in files if f.lower() not in ("readme.md", "notes.md")]
        if not readmes or not code_files:
            continue  # skip incomplete/in-progress folders

        meta = parse_readme(os.path.join(fpath, readmes[0]))
        if "title" not in meta:
            continue

        num_m = re.match(r"^(\d+)", folder)
        leetnum = int(num_m.group(1)) if num_m else None

        code_file = code_files[0]
        ext = code_file.split(".")[-1]
        lang_map = {"java": "Java", "py": "Python", "js": "JavaScript",
                    "cpp": "C++", "c": "C", "ts": "TypeScript"}
        with open(os.path.join(fpath, code_file), encoding="utf-8", errors="replace") as f:
            code = f.read()

        tags = fetch_tags(meta.get("slug"), cache) if meta.get("slug") else ["Uncategorized"]
        topic, subtopic = classify(tags)

        stats = git_commit_stats(folder)
        latest = stats.get("latest", {})
        date_str = stats.get("first_date")
        date_val = None
        if date_str:
            try:
                date_val = datetime.fromisoformat(date_str.replace("Z", "+00:00")).replace(tzinfo=None)
            except Exception:
                date_val = None

        records.append({
            "folder": folder,
            "leetnum": leetnum,
            "title": meta["title"],
            "difficulty": meta.get("difficulty", "Unknown"),
            "level": {"Easy": "Low", "Medium": "Medium", "Hard": "High"}.get(meta.get("difficulty"), "Unknown"),
            "url": meta.get("url"),
            "topic": topic,
            "subtopic": subtopic,
            "language": lang_map.get(ext, ext),
            "code": code,
            "runtime_ms": latest.get("runtime_ms"),
            "memory_mb": latest.get("memory_mb"),
            "attempts": stats.get("attempts", 1),
            "date_solved": date_val,
        })
    return records


# ----------------------------------------------------------------------------
# 3. Build the workbook (same structure as the manual version)
# ----------------------------------------------------------------------------
NAVY, NAVY_DARK, GOLD, GOLD_LIGHT = "1B2A4A", "0F1B33", "C9A24B", "F4E9CE"
WHITE, LIGHT_GREY = "FFFFFF", "F5F5F5"
EASY_GREEN, EASY_FILL = "2E7D32", "E8F5E9"
MED_ORANGE, MED_FILL = "B26A00", "FFF3E0"
HARD_RED, HARD_FILL = "C62828", "FDECEA"
FONT_NAME = "Calibri"

topic_font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
subtopic_font = Font(name=FONT_NAME, size=10.5, bold=True, color=NAVY_DARK)
header_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
normal_font = Font(name=FONT_NAME, size=10, color="222222")
topic_fill = PatternFill("solid", fgColor=NAVY)
subtopic_fill = PatternFill("solid", fgColor=GOLD_LIGHT)
header_fill = PatternFill("solid", fgColor=NAVY_DARK)
white_fill = PatternFill("solid", fgColor=WHITE)
grey_fill = PatternFill("solid", fgColor=LIGHT_GREY)
total_fill = PatternFill("solid", fgColor=GOLD)
thin = Side(style="thin", color="D9D9D9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_workbook(records):
    records = [r for r in records if r["title"]]
    records.sort(key=lambda r: (r["topic"], r["subtopic"],
                                 {"Low": 0, "Medium": 1, "High": 2}.get(r["level"], 3),
                                 r["leetnum"] or 0))

    by_topic_sub = defaultdict(lambda: defaultdict(list))
    for r in records:
        by_topic_sub[r["topic"]][r["subtopic"]].append(r)

    wb = Workbook()

    # ---- Sheet 1: Topic Summary ----
    ws1 = wb.active
    ws1.title = "Topic Summary"
    ws1.sheet_view.showGridLines = False
    cols1 = [("Topic", 16), ("Subtopic", 30), ("Low (Easy)", 12), ("Medium", 10),
             ("High (Hard)", 12), ("Total Solved", 12), ("Avg Runtime (ms)", 15), ("Avg Memory (MB)", 15)]
    for i, (h, w) in enumerate(cols1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.merge_cells("A1:H2")
    ws1["A1"] = "TOPIC-WISE SOLVE SUMMARY  (auto-updated)"
    ws1["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    ws1["A1"].fill = topic_fill
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for col in range(1, 9):
        ws1.cell(row=1, column=col).fill = topic_fill
        ws1.cell(row=2, column=col).fill = topic_fill

    hdr_r = 4
    for i, (h, w) in enumerate(cols1, start=1):
        c = ws1.cell(row=hdr_r, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, box
    ws1.freeze_panes = f"A{hdr_r + 1}"

    r = hdr_r + 1
    row_idx = 0
    for topic in sorted(by_topic_sub):
        first_sub = sorted(by_topic_sub[topic])[0]
        for subtopic in sorted(by_topic_sub[topic]):
            recs = by_topic_sub[topic][subtopic]
            low = sum(1 for x in recs if x["level"] == "Low")
            med = sum(1 for x in recs if x["level"] == "Medium")
            high = sum(1 for x in recs if x["level"] == "High")
            rts = [x["runtime_ms"] for x in recs if x.get("runtime_ms") is not None]
            mms = [x["memory_mb"] for x in recs if x.get("memory_mb") is not None]
            vals = [topic if subtopic == first_sub else "", subtopic, low, med, high, len(recs),
                    round(sum(rts) / len(rts), 1) if rts else None,
                    round(sum(mms) / len(mms), 1) if mms else None]
            fill = white_fill if row_idx % 2 == 0 else grey_fill
            for j, v in enumerate(vals, start=1):
                c = ws1.cell(row=r, column=j, value=v)
                c.font, c.fill, c.border = normal_font, fill, box
                c.alignment = left if j in (1, 2) else center
            r += 1
            row_idx += 1
        trecs = [x for sub in by_topic_sub[topic].values() for x in sub]
        low = sum(1 for x in trecs if x["level"] == "Low")
        med = sum(1 for x in trecs if x["level"] == "Medium")
        high = sum(1 for x in trecs if x["level"] == "High")
        vals = [f"{topic} — TOTAL", "", low, med, high, len(trecs), "", ""]
        for j, v in enumerate(vals, start=1):
            c = ws1.cell(row=r, column=j, value=v)
            c.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY_DARK)
            c.fill, c.border = subtopic_fill, box
            c.alignment = left if j in (1, 2) else center
        r += 1
        row_idx += 1

    low = sum(1 for x in records if x["level"] == "Low")
    med = sum(1 for x in records if x["level"] == "Medium")
    high = sum(1 for x in records if x["level"] == "High")
    vals = ["GRAND TOTAL", "", low, med, high, len(records), "", ""]
    for j, v in enumerate(vals, start=1):
        c = ws1.cell(row=r, column=j, value=v)
        c.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        c.fill, c.border = total_fill, box
        c.alignment = left if j in (1, 2) else center

    # ---- Sheet 2: All Problems (flat, filterable) ----
    ws2 = wb.create_sheet("All Problems")
    ws2.sheet_view.showGridLines = False
    cols2 = [("#", 4), ("LeetCode #", 10), ("Problem", 40), ("Topic", 12), ("Subtopic", 26),
             ("Difficulty", 10), ("Language", 9), ("Runtime (ms)", 11), ("Memory (MB)", 11),
             ("Attempts", 9), ("Date Solved", 12), ("Link", 8)]
    for i, (h, w) in enumerate(cols2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.merge_cells("A1:L2")
    ws2["A1"] = "ALL PROBLEMS — AUTO-SYNCED"
    ws2["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    ws2["A1"].fill = topic_fill
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for col in range(1, 13):
        ws2.cell(row=1, column=col).fill = topic_fill
        ws2.cell(row=2, column=col).fill = topic_fill
    hdr2 = 4
    for i, (h, w) in enumerate(cols2, start=1):
        c = ws2.cell(row=hdr2, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, box
    ws2.freeze_panes = f"A{hdr2 + 1}"

    records_by_date = sorted(records, key=lambda x: x["date_solved"] or datetime.min)
    r = hdr2 + 1
    for i, rec in enumerate(records_by_date):
        vals = [i + 1, rec["leetnum"], rec["title"], rec["topic"], rec["subtopic"],
                rec["difficulty"], rec["language"], rec.get("runtime_ms"), rec.get("memory_mb"),
                rec.get("attempts"), rec["date_solved"], "Open"]
        fill = white_fill if i % 2 == 0 else grey_fill
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(row=r, column=j, value=v)
            c.font, c.fill, c.border = normal_font, fill, box
            c.alignment = left if j in (3, 5) else center
        ws2.cell(row=r, column=11).number_format = "dd-mmm-yyyy"
        lc = ws2.cell(row=r, column=12, value="Open")
        lc.hyperlink = rec.get("url")
        lc.font = Font(name=FONT_NAME, size=10, color="1155CC", underline="single")
        r += 1
    last_r = r - 1
    ws2.conditional_formatting.add(f"F{hdr2+1}:F{last_r}", FormulaRule(
        formula=[f'F{hdr2+1}="Easy"'], fill=PatternFill("solid", fgColor=EASY_FILL),
        font=Font(color=EASY_GREEN, bold=True)))
    ws2.conditional_formatting.add(f"F{hdr2+1}:F{last_r}", FormulaRule(
        formula=[f'F{hdr2+1}="Medium"'], fill=PatternFill("solid", fgColor=MED_FILL),
        font=Font(color=MED_ORANGE, bold=True)))
    ws2.conditional_formatting.add(f"F{hdr2+1}:F{last_r}", FormulaRule(
        formula=[f'F{hdr2+1}="Hard"'], fill=PatternFill("solid", fgColor=HARD_FILL),
        font=Font(color=HARD_RED, bold=True)))

    # ---- Sheet 3: Topic Breakdown (Detailed) ----
    ws3 = wb.create_sheet("Topic Breakdown (Detailed)")
    ws3.sheet_view.showGridLines = False
    cols3 = [("", 3), ("Problem / Subtopic / Topic", 42), ("LeetCode #", 11), ("Difficulty", 10),
             ("Runtime (ms)", 12), ("Memory (MB)", 12), ("Language", 10), ("Date Solved", 13), ("Link", 8)]
    for i, (h, w) in enumerate(cols3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.merge_cells("A1:I2")
    ws3["A1"] = "TOPIC BREAKDOWN — DETAILED (auto-updated)"
    ws3["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    ws3["A1"].fill = topic_fill
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for col in range(1, 10):
        ws3.cell(row=1, column=col).fill = topic_fill
        ws3.cell(row=2, column=col).fill = topic_fill

    r = 4
    level_order = ["Low", "Medium", "High"]
    level_colors = {"Low": (EASY_FILL, EASY_GREEN), "Medium": (MED_FILL, MED_ORANGE), "High": (HARD_FILL, HARD_RED)}
    for topic in sorted(by_topic_sub):
        trecs = [x for sub in by_topic_sub[topic].values() for x in sub]
        low = sum(1 for x in trecs if x["level"] == "Low")
        med = sum(1 for x in trecs if x["level"] == "Medium")
        high = sum(1 for x in trecs if x["level"] == "High")
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws3.cell(row=r, column=1,
                      value=f"  TOPIC: {topic.upper()}   —   {len(trecs)} solved   (Low: {low} · Medium: {med} · High: {high})")
        c.font, c.fill = topic_font, topic_fill
        for col in range(1, 10):
            ws3.cell(row=r, column=col).fill = topic_fill
        r += 1
        for subtopic in sorted(by_topic_sub[topic]):
            recs = by_topic_sub[topic][subtopic]
            slow = sum(1 for x in recs if x["level"] == "Low")
            smed = sum(1 for x in recs if x["level"] == "Medium")
            shigh = sum(1 for x in recs if x["level"] == "High")
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            c = ws3.cell(row=r, column=1,
                          value=f"      Subtopic: {subtopic}   —   {len(recs)} solved   (Low: {slow} · Medium: {smed} · High: {shigh})")
            c.font, c.fill = subtopic_font, subtopic_fill
            for col in range(1, 10):
                ws3.cell(row=r, column=col).fill = subtopic_fill
            r += 1
            for level in level_order:
                level_recs = [x for x in recs if x["level"] == level]
                if not level_recs:
                    continue
                fillc, fontc = level_colors[level]
                ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
                label = {"Low": "Easy", "Medium": "Medium", "High": "Hard"}[level]
                c = ws3.cell(row=r, column=2, value=f"{level} ({label}) — {len(level_recs)} problem(s)")
                c.font = Font(name=FONT_NAME, size=9.5, bold=True, color=fontc)
                for col in range(1, 10):
                    ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fillc)
                r += 1
                for i, rec in enumerate(sorted(level_recs, key=lambda x: x["leetnum"] or 0)):
                    row_fill = white_fill if i % 2 == 0 else grey_fill
                    vals = ["", rec["title"], rec["leetnum"], rec["difficulty"],
                            rec.get("runtime_ms"), rec.get("memory_mb"), rec["language"],
                            rec["date_solved"], "Open"]
                    for j, v in enumerate(vals, start=1):
                        cc = ws3.cell(row=r, column=j, value=v)
                        cc.font, cc.fill, cc.border = normal_font, row_fill, box
                        cc.alignment = left if j == 2 else center
                    ws3.cell(row=r, column=8).number_format = "dd-mmm-yyyy"
                    lc = ws3.cell(row=r, column=9)
                    lc.hyperlink = rec.get("url")
                    lc.font = Font(name=FONT_NAME, size=10, color="1155CC", underline="single")
                    r += 1

    # ---- Sheet 4: Solutions ----
    ws4 = wb.create_sheet("Solutions")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 38
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 95
    ws4.merge_cells("A1:E2")
    ws4["A1"] = "SOLUTIONS — SOURCE CODE"
    ws4["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    ws4["A1"].fill = topic_fill
    ws4["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for col in range(1, 6):
        ws4.cell(row=1, column=col).fill = topic_fill
        ws4.cell(row=2, column=col).fill = topic_fill
    hdr4 = 4
    for i, h in enumerate(["#", "Problem", "LeetCode #", "Language", "Code"], start=1):
        c = ws4.cell(row=hdr4, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, box
    ws4.freeze_panes = f"A{hdr4 + 1}"
    mono_font = Font(name="Consolas", size=9, color="1B2A4A")
    r = hdr4 + 1
    for i, rec in enumerate(sorted(records, key=lambda x: x["leetnum"] or 0)):
        code = rec.get("code") or ""
        n_lines = code.count("\n") + 1
        ws4.row_dimensions[r].height = min(max(14 * min(n_lines, 40), 15), 560)
        vals = [i + 1, rec["title"], rec["leetnum"], rec["language"], code]
        fill = white_fill if i % 2 == 0 else grey_fill
        for j, v in enumerate(vals, start=1):
            c = ws4.cell(row=r, column=j, value=v)
            c.fill = fill
            c.border = box
            if j == 5:
                c.font = mono_font
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.font = normal_font
                c.alignment = center if j != 2 else left
        r += 1

    return wb


def main():
    print("Scanning repo and fetching official LeetCode tags...")
    records = collect_records()
    print(f"Found {len(records)} solved problems.")
    wb = build_workbook(records)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
